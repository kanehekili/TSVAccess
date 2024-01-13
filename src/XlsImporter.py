'''
Created on Dec 6, 2023
Module for importing Xls sheets into the database (conplan) 
@author: Kanehekili
'''
import traceback,os,argparse,sys,json
from PyQt6.QtWidgets import QApplication, QTextEdit
from PyQt6.QtCore import Qt,pyqtSignal
from PyQt6 import QtWidgets, QtGui, QtCore
#install python-openpyxl (arch native) or pylightxl which handles Softmaker better
from openpyxl import load_workbook
from DBTools import OSTools
from TsvDBCreator import SetUpTSVDB
import DBTools
from datetime import datetime
from time import sleep
from enum import Enum
from collections import Counter



'''
Unbelievable windows crap: To get your icon into the task bar:
'''
if os.name == 'nt':
    import ctypes
    myappid = 'import.tsv.access'  # arbitrary string
    cwdll = ctypes.windll  # @UndefinedVariable
    cwdll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

# #Main App Window. 
class MainFrame(QtWidgets.QMainWindow):
    
    def __init__(self, qapp):
        self._isStarted = False
        self.__qapp = qapp
        self.qtQueueRunning = False
        self.sourceFile=None
        self.converter=Converter()
        super(MainFrame, self).__init__()
        self.setWindowIcon(getAppIcon())
        self.initUI()
        self.centerWindow()
        self.setWindowTitle("TSV XLSX Importer")
        self.show()
        qapp.applicationStateChanged.connect(self.__queueStarted)    

    def centerWindow(self):
        frameGm = self.frameGeometry()
        centerPoint = self.screen().availableGeometry().center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def __queueStarted(self, _state):
        if self.qtQueueRunning:
            return
        self.qtQueueRunning = True
        self._initModel()
        
    def initUI(self):
        #box = self.makeGridLayout()
        self.init_toolbar()
        self.uiInfoLabel=QTextEdit(self)
        self.uiInfoLabel.setReadOnly(True)
        self.uiInfoLabel.setAcceptRichText(True)
        self.uiInfoLabel.setText("<h3>Conplan Importer<h3> ..mit Datenbank verbinden")

        self.uiInfoLabel.setAlignment(Qt.AlignmentFlag.AlignTop)
        box = QtWidgets.QVBoxLayout();
        box.addWidget(self.uiInfoLabel)
        
        # Without central widget it won't work
        wid = QtWidgets.QWidget(self)
        self.setCentralWidget(wid)    
        # TODO: self.resize
        wid.setLayout(box)
        self.resize(400, 450)


    def init_toolbar(self):
        self.startAction = QtGui.QAction(QtGui.QIcon('./web/static/Button-Download.png'), 'Datei öffnen', self)
        self.startAction.setShortcut('Ctrl+M')
        self.startAction.triggered.connect(self.getSourceFile)

        #self.importAction = QtGui.QAction(QtGui.QIcon('./web/static/save.png'), "Importieren", self)
        #self.importAction.setShortcut('Ctrl+T')
        #self.importAction.triggered.connect(self.startImport)



        self.stopAction = QtGui.QAction(QtGui.QIcon('./web/static/dialog-close.png'), 'Schließen', self)
        self.stopAction.setShortcut('Ctrl+H')
        self.stopAction.triggered.connect(self.goodbye)
        #self.stopAction.setEnabled(False)
        
        spacer = QtWidgets.QWidget();
        spacer.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred);
        
        self.toolbar = self.addToolBar('Main')
        self.toolbar.addAction(self.startAction)
        self.toolbar.addSeparator()
        self.toolbar.addWidget(spacer);
        self.toolbar.addAction(self.stopAction)
        

    
    def _initModel(self):
        QApplication.setOverrideCursor(QtCore.Qt.CursorShape.WaitCursor)
        res = self.converter.connectDB()
        QApplication.restoreOverrideCursor()        
        if not res:
            dlg = self.getErrorDialog("Datenbank Fehler ", "Datenbank nicht gefunden", "Es besteht keine Verbindung zur Datenbank. Bitte melden!")
            dlg.buttonClicked.connect(self._onErrorDialogClicked)
            dlg.show()
        self.uiInfoLabel.setHtml(self.converter.exportInstructions())

    def _onErrorDialogClicked(self, _):
        self.goodbye()
    
    def enableActions(self,enable):
        self.stopAction.setEnabled(enable)
        self.startAction.setEnabled(enable)
    
    @QtCore.pyqtSlot()
    def getSourceFile(self):
        targetPath = self.readLastXPath()
        extn = "*.xlsx *.xls"
        fileFilter = "Spreadsheets (%s);;Alle Dateien(*.*)" % extn
        
        result = QtWidgets.QFileDialog.getOpenFileName(parent=self, directory=targetPath, caption="Importiere Mitgliederdaten",filter=fileFilter);
        if result[0]:
            self.sourceFile=result[0]
            self.saveLastXPath(self.sourceFile)
            QApplication.setOverrideCursor(QtCore.Qt.CursorShape.WaitCursor)
            worker = LongRunningOperation(self._startImport,self.sourceFile)
            worker.signal.connect(self._cleanupWorker)
            worker.startOperation()  
 
    def _startImport(self,path):
        print("Importing path:",path)
        self.enableActions(False)
        fullsheet= XImporter.importXel(path)
        self.converter.data=fullsheet
        res=self.converter.verifySheet()
        if res:
            mbrList=self.converter.buildModel()
            #self.uiInfoLabel.setHtml("<b>Fertig</b>")
            self.converter.persistCSV(mbrList)
            
    
    def _cleanupWorker(self, worker):
        # QThread: Destroyed while thread is still running
        msg = worker.msg 
        print("Long operation done:%s",msg)
        self.enableActions(True)
        if msg:
            dlg=self.getErrorDialog("Fehler", msg, "Im log kann man am besten sehen, was nicht funktioniert hat.")
            dlg.show()
        else:
            self.uiInfoLabel.setHtml(self.converter.htmlReport())
        worker.quit()
        worker.wait(); 
        QApplication.restoreOverrideCursor() 
    
    @QtCore.pyqtSlot()    
    def goodbye(self):
        self.close()
    
    def getErrorDialog(self, text, infoText, detailedText):
        dlg = QtWidgets.QMessageBox(self)
        dlg.setIcon(QtWidgets.QMessageBox.Icon.Warning)
        dlg.setWindowModality(Qt.WindowModality.WindowModal)
        dlg.setWindowTitle("Fehler")
        dlg.setText(text)
        dlg.setInformativeText(infoText)
        dlg.setDetailedText(detailedText)
        dlg.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok)
        spacer = QtWidgets.QSpacerItem(300, 1, QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Minimum)
        layout = dlg.layout()
        layout.addItem(spacer, layout.rowCount(), 0, 1, layout.columnCount())
        #msg = infoText + "\n DETAIL:" + detailedText
        return dlg    

    def readLastXPath(self):
        cfg = self._configPath()
        if OSTools.fileExists(cfg):
            with open(cfg) as jsonFile:
                data = json.load(jsonFile)
                return data["SRC"]
        return "."
    def saveLastXPath(self,path):
        data={}
        cfg = self._configPath()
        data["SRC"]=path
        with open(cfg, 'w') as jsonFile:
            json.dump(data, jsonFile)

    def _configPath(self):
        folder= OSTools.joinPathes(OSTools().getHomeDirectory(),".config")
        OSTools.ensureDirectory(folder)
        return OSTools.joinPathes(folder,"xlsimporter.json")         

''' Long running operations for actions that do not draw or paint '''        
class LongRunningOperation(QtCore.QThread):
    signal = pyqtSignal(object) 

    def __init__(self, func, *args):
        QtCore.QThread.__init__(self)
        self.function = func
        self.arguments = args
        self.msg=None

    def run(self):
        try:
            self.function(*self.arguments)
        except Exception as ex:
            Log.exception("***Error in LongRunningOperation***")
            self.msg = "Problem: "+str(ex)
        finally:
            self.signal.emit(self)

    def startOperation(self):
        self.start()  # invokes run - process pending QT events
        sleep(0.5)
        QtCore.QCoreApplication.processEvents()


class XImporter:
    def __init__(self):
        pass
    
    #only importer knowlege of xlsx files. returns array of rowArrays=cell
    @classmethod
    def importXel(cls,path):
        wb = load_workbook(filename=path, read_only=True)
        sheets= wb.sheetnames
        ws=wb[sheets[0]]
        fullSheet=[]
        for row in ws.rows:
            singleRow=[]
            for cell in row:
                saveVal=cell.value if cell.value else '' 
                singleRow.append(saveVal)
            fullSheet.append(singleRow)
        cntMbrs=len(fullSheet)
        cntCols = len(fullSheet[0])
        Log.info("Read %d entries, with %d columns each",cntMbrs,cntCols)
        wb.close()        
        return fullSheet

class Fields(Enum):
    ID = 0
    GENDER = 1
    NAME = 2
    VORNAME = 3
    ACCESS = 4
    EOLDATE = 5 #not relevant
    BDATE = 6
    SECDATE=7
    SECTION=8

class TsvMember():
    PAYOK="-"
    def __init__(self,cpid,fn,ln,access,gender,birthdate):
        self.baseData=(cpid,fn,ln,access,gender,birthdate)
        self.payData={} # section-> paydate. output must be array of tuple(secion,paydate)
    
    def getID(self):
        return self.baseData[0]
    
    def getName(self):        
        return self.baseData[2]
    
    def getAccess(self):
        return self.baseData[3]

    #if that section is already there and with empty pay - it it invalid - latest pay counts, empty rules
    def addPay(self,payDate,section):
        pd = self.payData.get(section,None)
        if not pd:
            self.payData[section]=payDate
            return
        #tricky
        if pd == TsvMember.PAYOK:
            return #payok rules
        
        if pd < payDate:
            self.payData[section]=payDate            
    
    def sectionData(self):
        data=[]
        for key,pd in self.payData.items():
            if pd == TsvMember.PAYOK:
                pd=None 
            data.append((self.getID(),pd,key))
        return data
     
     
    def sections(self):
        return self.payData.keys()    
    
    def display(self):
        print(self.baseData,">",self.payData)


class Converter():
    ROW_SIZE=9
    def __init__(self,fullSheet=None):
        self.data=fullSheet # array of array(row-cells)
        self.findingsImport=["<h2> Kein Erfolg </h2>"]
        self.dbSystem=None

    def connectDB(self):
        self.dbSystem = SetUpTSVDB("TsvDB")#TODO must be executed at start - incl error dialog
        return self.dbSystem.isConnected()

    def exportInstructions(self):
        txt=[]
        txt.append("<h3>Hilfe Conplan Export</h3>")
        txt.append("<ul><li>Personendaten -> Listengenerator (Menue)</li>")
        txt.append("<li>Suchmodus ändern <strong>[click]</strong></li>")
        txt.append("<li>Adressen-Beitragszuordnung.. übernehmen <strong>[click]</strong></li>")
        txt.append("<li> Suchkriterium -> aktuelle Mitglieder suchen <strong><wählen></strong></li>")
        txt.append("<li>Suche ausführen <strong>[click]</strong></li>")
        txt.append("<li>Daten exportieren (Menue)</li>")
        txt.append("<li>Ausgabeprofil (unten) -> TsvAccessBeitrag laden <strong>[click]</strong></li>")
        txt.append("<li>Export fortsetzen <strong>[click]</strong></li>")
        txt.append("<li>Export starten <strong>[click]</strong></li>")
        txt.append('<li>Im Dialog "Herunterladen" + beenden</li>')
        txt.append("<li><i>Warten bis der Brauser die heruntergeladene Datei anzeigt</i></li>")
        txt.append("</ul>")
        txt.append("Viel Erfolg")
        return ''.join(txt)
        

    def htmlReport(self):
        return ''.join(self.findingsImport)

    def verifySheet(self):
        if len(self.data)<100:
            self.findingsImport.append("<br>Nicht genügend Einträge im Datensatz")
            return False
        if len(self.data[0])!=Converter.ROW_SIZE:
            self.findingsImport.append("<br>Fehler bei der Anzahl der Spalten")
            return False
        return True
    
    def buildModel(self):
        data = {} #dic of id-> TsvMember
        multiSet = {}
        sections=[]
        rogue=[]
        mbrCount=0    
        for line in self.data:
            # print(line)
            isHeader = "Vorname" in line
            if isHeader:
                continue
            conplanID=line[Fields.ID.value]
            mbr = data.get(conplanID,None)
            if not mbr:
                fn = line[Fields.VORNAME.value]
                nn = line[Fields.NAME.value]
                access = ''
                zugang = line[Fields.ACCESS.value].split(' ')
                if len(zugang) >= 1:
                    access = zugang[0].upper()
            
                tmpGender = line[Fields.GENDER.value]
                if tmpGender.startswith("w"):
                    gender = "F"
                else:
                    gender = "M"
                
                tmpDate = line[Fields.BDATE.value]
                if len(tmpDate) > 0:
                    birthdate = datetime.strptime(tmpDate, '%d.%m.%Y').date().isoformat()
                else:
                    birthdate = None

                if multiSet.get(access, None) is None:
                    multiSet[access] = 1
                else:
                    multiSet[access] += 1
                #cpid,fn,ln,access,gender,birthdate
                mbr=TsvMember(conplanID,fn,nn,access,gender,birthdate)
                data[conplanID]=mbr
                mbrCount += 1
                
            #--done -now the section data
            tmpDate=line[Fields.SECDATE.value]
            if len(tmpDate) > 0:
                payDate = datetime.strptime(tmpDate, '%d.%m.%Y').date().isoformat()
            else:
                payDate= TsvMember.PAYOK
                
            section= line[Fields.SECTION.value] #not empty
            mbr.addPay(payDate,section)                
                
        for entry in data.values():
            #entry.display()
            xxxsec=entry.sections()
            sections.extend(xxxsec)
            if "Hauptverein" not in xxxsec:
                rogue.append((entry.getID(),entry.getName()))
            
        self.makeImportFindings(sections,multiSet, mbrCount,rogue) 
        return list(data.values()) #array of TsvMembers        

    def persistCSV(self,memberList):
        txt=[]
        txt.append("<h2>Änderungen</h2>")
        
        try:
            txt.append("<ul>")
            lostMembers=self.symDiff(memberList,txt)
            for pk in lostMembers:
                stmt="UPDATE Mitglieder set flag=1 where id=%d"%(pk)
                self.dbSystem.db.select(stmt)
            txt.append("<li>%d Mitglieder ausgetreten</li>"%(len(lostMembers)))
            self.dbSystem.updateMembers(memberList)
        except Exception:
            #traceback.print_exc()
            Log.exception("Persistence failed")
            raise Exception("Fehler beim importieren!")
        finally:
            txt.append("</ul>")
            self.findingsImport.extend(txt)
            self.dbSystem.close()

    def symDiff(self,importData,txtBuffer):
        stmt="select id,flag from %s"%(SetUpTSVDB.MAINTABLE)
        rows=self.dbSystem.db.select(stmt)
        ids = [data[0] for data in rows] #int
        currIds=[int(mbr.getID()) for mbr in importData]
        txtBuffer.append("<li>%d Mitglieder werden übertragen</li>"%(len(currIds)))
        diff=[]
        for idFlag in rows:
            if not idFlag[0] in currIds:
                if idFlag[1] == 0: #not flagged yet
                    #print("!Member lost:%d -will be flagged!"%(idFlag[0]))
                    Log.info("Member lost:%d",idFlag[0])
                    diff.append(idFlag[0])     
        #the other way:
        newCount=0
        for cid in currIds:
            if not cid in ids:
                newCount+=1
                Log.info("%d) New Member:%d",newCount,cid)
        txtBuffer.append("<li>%d neue Mitgleider übernommen</li>"%(newCount))
        return diff    
    
    def makeImportFindings(self, sections,multiSet, mbrCount,rogue):
        txt=[]
        txt.append("<h2>Import Statistik</h2>")
        txt.append("<ul><li>Importiert:%d</li>" % (mbrCount))
        for key, cnt in multiSet.items():
            txt.append("<li>%s:%d</li>"%(key,cnt))
        txt.append("<li>Nicht im Hauptverein:%d</li>"%(len(rogue)))
        txt.append("<li>Statistik Abeteilungen:</li>")
        c = Counter(sections)
        txt.append("<ul>")
        for entry in c:
            txt.append("<li>%s %d</li>"%(entry,c[entry]))
        txt.append("</ul>")
        txt.append("</ul>")
        self.findingsImport=txt


def headlessImport(path):
    print("Importing path headless:",path)
    fullsheet= XImporter.importXel(path)
    c=Converter(fullsheet)
    res=c.verifySheet()
    if res:
        mbrList=c.buildModel()
        c.persistCSV(mbrList)
        
    print(c.findingsImport)    
        
    
def getAppIcon():
    return QtGui.QIcon('./web/static/TSV-import.png')


def handle_exception(exc_type, exc_value, exc_traceback):
    """ handle all exceptions """
    if WIN is not None:
        infoText = str(exc_value)
        detailText = "*".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
        WIN.getErrorDialog("Unexpected error", infoText, detailText).show()
        Log.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))


def parse():
    parser = argparse.ArgumentParser(description="XlsImporter")
    parser.add_argument('-d', dest="debug", action='store_true', help="Debug logs")
    parser.add_argument('-f', dest="headless", type=str, default=None, help="execute without GUI with filename",metavar="FILE")
    # parser.add_argument('-r', dest="rfidMode", action='store_true', help="RFID MODE")
    return parser.parse_args()    


def main(args):

    # todo speed up camera lookup. check log and promote: -c nbr
    #if OSTools.checkIfInstanceRunning("XlsImporter"):
    #    print("App already running")
    #    sys.exit()
        
    try:
        global Log
        global WIN
        wd = OSTools.getLocalPath(__file__)
        OSTools.setMainWorkDir(wd)
        OSTools.setupRotatingLogger("TSVXLS", args.debug)
        Log = DBTools.Log
        if args.debug:
            OSTools.setLogLevel("Debug")
        else:
            OSTools.setLogLevel("Info")
        fn = args.headless
        if fn is not None:
            headlessImport(fn)
            return 0
        Log.info("--- Start Xls Importer ---")
        argv = sys.argv
        app = QApplication(argv)
        app.setWindowIcon(getAppIcon())
        WIN = MainFrame(app)  # keep python reference!
        # ONLY windoze, if ever: app.setStyleSheet(winStyle())
        # app.setStyle(QtWidgets.QStyleFactory.create("Fusion"));
        app.exec()
        # logging.shutdown()
    except:
        with open('/tmp/error.log','a') as f:
            f.write(traceback.format_exc())
        traceback.print_exc(file=sys.stdout)
        Log.exception("Error in main:")
        # ex_type, ex_value, ex_traceback
        sys_tuple = sys.exc_info()
        QtWidgets.QMessageBox.critical(None, "Error!", str(sys_tuple[1]))


if __name__ == '__main__':
    sys.excepthook = handle_exception
    sys.exit(main(parse()))
    pass